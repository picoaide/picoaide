"""工具执行器：根据权限执行 MCP 工具调用"""

import base64
import io
import json
import os
import logging
import platform

import pyautogui
from PIL import Image
from mss import mss

from .permissions import is_tool_allowed

logger = logging.getLogger(__name__)

# 安全设置
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.05


def check_whitelist(path, whitelist_dirs):
  """检查路径是否在白名单内"""
  if not whitelist_dirs:
    return False
  abs_path = os.path.abspath(path)
  for d in whitelist_dirs:
    if abs_path.startswith(os.path.abspath(d)):
      return True
  return False


def execute_tool(tool, params, permissions, whitelist_dirs):
  """执行工具调用，返回结果字典"""
  if not is_tool_allowed(tool, permissions):
    return {"error": {"message": f"工具 {tool} 未被授权"}}

  # 文件类工具需要白名单检查
  if tool in ("computer_file_read", "computer_file_write", "computer_file_list", "computer_file_search"):
    path = params.get("path", "")
    if not check_whitelist(path, whitelist_dirs):
      return {"error": {"message": f"路径不在白名单内: {path}"}}

  try:
    # computer_whitelist 特殊处理：直接返回白名单目录
    if tool == "computer_whitelist":
      return {"result": {"directories": whitelist_dirs, "count": len(whitelist_dirs)}}

    handler = TOOL_HANDLERS.get(tool)
    if handler is None:
      return {"error": {"message": f"不支持的工具: {tool}"}}
    result = handler(params)
    return {"result": result}
  except Exception as e:
    logger.exception("工具执行失败: %s", tool)
    return {"error": {"message": str(e)}}


# ============================================================
# 工具实现
# ============================================================

def _screenshot(params):
  with mss() as sct:
    # grab() 直接返回 PIL 兼容的截图数据，不写文件
    monitor = sct.monitors[0]
    shot = sct.grab(monitor)
    img = Image.frombytes("RGB", shot.size, shot.bgra, "raw", "BGRX")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode()
  return {
    "content": [
      {"type": "image", "data": b64, "mimeType": "image/png"}
    ]
  }


def _screen_size(params):
  w, h = pyautogui.size()
  return {"width": w, "height": h, "platform": platform.system()}


def _active_window(params):
  """获取当前活跃窗口的标题"""
  import subprocess
  system = platform.system()
  try:
    if system == "Windows":
      import ctypes
      hwnd = ctypes.windll.user32.GetForegroundWindow()
      length = ctypes.windll.user32.GetWindowTextLengthW(hwnd) + 1
      buf = ctypes.create_unicode_buffer(length)
      ctypes.windll.user32.GetWindowTextW(hwnd, buf, length)
      title = buf.value
    elif system == "Darwin":
      result = subprocess.run(
        ["osascript", "-e", 'tell application "System Events" to get name of first process whose frontmost is true'],
        capture_output=True, text=True, timeout=5
      )
      title = result.stdout.strip()
    else:
      result = subprocess.run(["xdotool", "getactivewindow", "getwindowname"],
        capture_output=True, text=True, timeout=5)
      title = result.stdout.strip()
    return {"title": title, "platform": system}
  except Exception as e:
    return {"title": "", "error": str(e)}


def _mouse_click(params):
  x, y = int(params["x"]), int(params["y"])
  button = params.get("button", "left")
  clicks = int(params.get("clicks", 1))
  pyautogui.click(x=x, y=y, button=button, clicks=clicks)
  return {"ok": True, "x": x, "y": y}


def _mouse_move(params):
  x, y = int(params["x"]), int(params["y"])
  pyautogui.moveTo(x, y)
  return {"ok": True}


def _mouse_drag(params):
  sx, sy = int(params["startX"]), int(params["startY"])
  ex, ey = int(params["endX"]), int(params["endY"])
  pyautogui.moveTo(sx, sy)
  pyautogui.drag(ex - sx, ey - sy, duration=0.3)
  return {"ok": True}


def _mouse_scroll(params):
  x = int(params.get("x", 0))
  y = int(params.get("y", 0))
  sx = int(params.get("scrollX", 0))
  sy = int(params.get("scrollY", 0))
  if x or y:
    pyautogui.moveTo(x, y)
  pyautogui.hscroll(sx)
  pyautogui.scroll(sy)
  return {"ok": True}


def _keyboard_type(params):
  text = params["text"]
  pyautogui.write(text, interval=0.02)
  return {"ok": True}


def _keyboard_press(params):
  keys = params["keys"]
  if isinstance(keys, list) and len(keys) > 1:
    pyautogui.hotkey(*keys)
  elif isinstance(keys, list) and len(keys) == 1:
    pyautogui.press(keys[0])
  else:
    pyautogui.press(str(keys))
  return {"ok": True}


def _file_read(params):
  path = params["path"]
  if not os.path.isfile(path):
    return {"error": {"message": f"文件不存在: {path}"}}

  ext = os.path.splitext(path)[1].lower()

  # xlsx
  if ext == ".xlsx":
    return _read_xlsx(path)
  # docx
  if ext == ".docx":
    return _read_docx(path)
  # 普通文本
  try:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
      content = f.read()
    return {"content": [{"type": "text", "text": content}], "path": path}
  except Exception as e:
    return {"error": {"message": str(e)}}


def _read_xlsx(path):
  import openpyxl
  wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
  sheets_data = {}
  for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
      rows.append([str(c) if c is not None else "" for c in row])
    sheets_data[name] = rows
  wb.close()
  return {
    "content": [{"type": "text", "text": json.dumps(sheets_data, ensure_ascii=False)}],
    "path": path,
    "sheets": list(sheets_data.keys()),
  }


def _read_docx(path):
  from docx import Document
  doc = Document(path)
  paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
  return {
    "content": [{"type": "text", "text": "\n".join(paragraphs)}],
    "path": path,
  }


def _file_write(params):
  path = params["path"]
  content = params["content"]
  os.makedirs(os.path.dirname(path), exist_ok=True)
  with open(path, "w", encoding="utf-8") as f:
    f.write(content)
  return {"ok": True, "path": path}


def _file_list(params):
  path = params["path"]
  if not os.path.isdir(path):
    return {"error": {"message": f"目录不存在: {path}"}}
  entries = []
  for entry in sorted(os.scandir(path), key=lambda e: (not e.is_dir(), e.name.lower())):
    try:
      stat = entry.stat()
      entries.append({
        "name": entry.name,
        "is_dir": entry.is_dir(),
        "size": stat.st_size,
      })
    except OSError:
      continue
  return {"entries": entries, "path": path}


def _whitelist(params):
  """白名单目录列表 — 需要外部传入"""
  # 这个函数在 execute_tool 中特殊处理
  return {}


def _file_search(params):
  path = params["path"]
  query = params.get("query", "").lower()
  max_results = int(params.get("max_results", 50))
  if not os.path.isdir(path):
    return {"error": {"message": f"目录不存在: {path}"}}
  matches = []
  for root, dirs, files in os.walk(path):
    for name in files + dirs:
      if query in name.lower():
        full = os.path.join(root, name)
        try:
          is_dir = os.path.isdir(full)
          size = os.path.getsize(full) if not is_dir else 0
          matches.append({"path": full, "name": name, "is_dir": is_dir, "size": size})
          if len(matches) >= max_results:
            break
        except OSError:
          continue
    if len(matches) >= max_results:
      break
  return {"matches": matches, "count": len(matches), "query": query}


# 工具名 -> 处理函数映射
TOOL_HANDLERS = {
  "computer_screenshot": _screenshot,
  "computer_screen_size": _screen_size,
  "computer_active_window": _active_window,
  "computer_mouse_click": _mouse_click,
  "computer_mouse_move": _mouse_move,
  "computer_mouse_drag": _mouse_drag,
  "computer_mouse_scroll": _mouse_scroll,
  "computer_keyboard_type": _keyboard_type,
  "computer_keyboard_press": _keyboard_press,
  "computer_file_read": _file_read,
  "computer_file_write": _file_write,
  "computer_file_list": _file_list,
  "computer_whitelist": _whitelist,
  "computer_file_search": _file_search,
}
